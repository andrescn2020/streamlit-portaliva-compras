import streamlit as st
import pandas as pd
import zipfile
import io
import re
import openpyxl
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side

st.title("Lector de CSV desde archivo ZIP")

# Diccionario de equivalencias de códigos a nombres de comprobantes
comprobante_dict = {
    "001": "FACTURA A",
    "002": "NOTA DE DEBITO A",
    "003": "NOTA DE CREDITO A",
    "004": "RECIBO A",
    "005": "NOTA DE VENTA AL CONTADO A",
    "006": "FACTURA B",
    "007": "NOTA DE DEBITO B",
    "008": "NOTA DE CREDITO B",
    "009": "RECIBO B",
    "010": "NOTA DE VENTA AL CONTADO B",
    "011": "FACTURA C",
    "012": "NOTA DE DEBITO C",
    "013": "NOTA DE CREDITO C",
    "014": "DOCUMENTO ADUANERO",
    "015": "RECIBO C",
    "016": "NOTA DE VENTA AL CONTADO C",
    "017": "FACTURA DE EXPORTACION",
    "018": "NOTA DE DEBITO POR OPERACIONES CON EL EXTERIOR",
    "019": "NOTA DE CREDITO POR OPERACIONES CON EL EXTERIOR",
    "020": "RECIBO POR OPERACIONES CON EL EXTERIOR",
    "021": "FACTURA - PERMISO EXPORTACION SIMPLIFICADO - DTO. 855/97",
    "022": "COMPROBANTE DE COMPRA DE BIENES USADOS",
    "023": "MANDATO - CONSIGNACION",
    "024": "COMPROBANTE PARA RECICLAR MATERIALES",
    "034": "COMPROBANTE A DEL APARTADO A  INCISO F  R G N° 1415",
    "035": "COMPROBANTE B DEL ANEXO I, APARTADO A, INC.F), RG N° 1415",
    "036": "COMPROBANTE C DEL Anexo I, Apartado A, INC.F), R.G. N° 1415",
    "037": "NOTA DE DEBITO O DOCUMENTO EQUIVALENTE QUE CUMPLAN CON LA R.G. N° 1415",
    "038": "NOTA DE CREDITO O DOCUMENTO EQUIVALENTE QUE CUMPLAN CON LA R.G. N° 1415",
    "039": "OTRO COMPROBANTES A QUE CUMPLAN CON LA R.G. N° 1415",
    "040": "OTRO COMPROBANTES B QUE CUMPLAN CON LA R.G. N° 1415",
    "041": "OTRO COMPROBANTES C QUE CUMPLAN CON LA R.G. N° 1415",
    "043": "RECIBO FACTURA A  REGIMEN DE FACTURA DE CREDITO",
    "051": "FACTURA M",
    "052": "NOTA DE DEBITO M",
    "053": "NOTA DE CREDITO M",
    "054": "RECIBO M",
    "055": "NOTA DE VENTA AL CONTADO M",
    "056": "COMPROBANTES M DEL ANEXO I APARTADO A  INC F  R G N 1415",
    "057": "OTRO COMPROBANTES M QUE CUMPLAN CON LA R G N 1415",
    "058": "CUENTA DE VENTA Y LIQUIDO PRODUCTO A",
    "059": "LIQUIDACIONES M",
    "060": "CUENTA DE VENTA Y LIQUIDO PRODUCTO B",
    "061": "CUENTA DE VENTA Y LIQUIDO PRODUCTO C",
    "063": "LIQUIDACIONE A",
    "064": "LIQUIDACIONE B",
    "065": "NOTA DE CREDITO DE COMPROBANTES CON COD. 34, 39, 58, 59, 60, 63, 96, 97",
    "066": "DESPACHO DE IMPORTACION",
    "067": "IMPORTACION DE SERVICIOS",
    "068": "LIQUIDACION C",
    "069": "RECIBO FACTURA DE CREDITO",
    "070": "CREDITO FISCAL POR CONTRIBUCIONES PATRONALES",
    "071": "FORMULARIO 1116 RT",
    "074": "CARTA DE PORTE PARA EL TRANSPORTE AUTOMOTOR PARA GRANOS",
    "075": "CARTA DE PORTE PARA EL TRANSPORTE FERROVIARIO PARA GRANOS",
    "080": "COMPROBANTE DIARIO DE CIERRE (ZETA)",
    "081": "TIQUE FACTURA A  CONTROLADORES FISCALES",
    "082": "TIQUE - FACTURA B",
    "083": "TIQUE",
    "084": "COMPROBANTE  FACTURA DE SERVICIOS PUBLICOS  INTERESES FINANCIEROS",
    "085": "NOTA DE CREDITO  SERVICIOS PUBLICOS  NOTA DE CREDITO CONTROLADORES FISCALES",
    "086": "NOTA DE DEBITO  SERVICIOS PUBLICOS",
    "087": "OTROS COMPROBANTES - SERVICIOS DEL EXTERIOR",
    "088": "OTROS COMPROBANTES - DOCUMENTOS EXCEPTUADOS / REMITO ELECTRONICO",
    "089": "OTROS COMPROBANTES - DOCUMENTOS EXCEPTUADOS - NOTAS DE DEBITO / RESUMEN DE DATOS",
    "090": "OTROS COMPROBANTES - DOCUMENTOS EXCEPTUADOS - NOTAS DE CREDITO",
    "091": "REMITO R",
    "092": "AJUSTES CONTABLES QUE INCREMENTAN EL DEBITO FISCAL",
    "093": "AJUSTES CONTABLES QUE DISMINUYEN EL DEBITO FISCAL",
    "094": "AJUSTES CONTABLES QUE INCREMENTAN EL CREDITO FISCAL",
    "095": "AJUSTES CONTABLES QUE DISMINUYEN EL CREDITO FISCAL",
    "096": "FORMULARIO 1116 B",
    "097": "FORMULARIO 1116 C",
    "099": "OTROS COMP QUE NO CUMPLEN CON LA R G 3419 Y SUS MODIF",
    "100": "AJUSTE ANUAL PROVENIENTE DE LA D J DEL IVA POSITIVO",
    "101": "AJUSTE ANUAL PROVENIENTE DE LA D J DEL IVA NEGATIVO",
    "102": "NOTA DE ASIGNACION",
    "103": "NOTA DE CREDITO DE ASIGNACION",
    "104": "NOTA DE DEBITO DE ASIGNACION",
}

# Campo obligatorio para el nombre del contribuyente (primero)
contribuyente = st.text_input(
    "Nombre del contribuyente (obligatorio)", value="", max_chars=100
)
contribuyente = contribuyente.strip().upper()
if not contribuyente:
    st.error("Debes ingresar el nombre del contribuyente.")
    st.stop()

# Luego el selector de archivo ZIP
uploaded_file = st.file_uploader(
    "Selecciona un archivo ZIP que contenga un CSV", type=["zip"]
)

if uploaded_file is not None:
    try:
        # Leer el archivo ZIP
        with zipfile.ZipFile(uploaded_file) as zip_file:
            # Obtener la lista de archivos en el ZIP
            file_list = zip_file.namelist()

            # Buscar archivos CSV en el ZIP
            csv_files = [f for f in file_list if f.endswith(".csv")]

            if csv_files:
                # Leer el primer archivo CSV encontrado
                with zip_file.open(csv_files[0]) as csv_file:
                    # Intentar diferentes opciones de lectura
                    try:
                        # Primero intentamos con el delimitador por defecto (coma)
                        df = pd.read_csv(csv_file, encoding="utf-8")
                    except:
                        try:
                            # Si falla, intentamos con punto y coma
                            csv_file.seek(0)
                            df = pd.read_csv(csv_file, sep=";", encoding="utf-8")
                        except:
                            # Si aún falla, intentamos con tab
                            csv_file.seek(0)
                            df = pd.read_csv(csv_file, sep="\t", encoding="utf-8")

                    # Renombrar columnas
                    column_rename = {
                        "Importe de Percepciones de Ingresos Brutos": "Perc. IIBB",
                        "Importe de Percepciones o Pagos a Cuenta de IVA": "Perc. IVA",
                        "Importe Exento": "Exento",
                        "Importe No Gravado": "No Gravado",
                        "Fecha de Emisión": "Fecha",
                        "Tipo de Comprobante": "Comprobante",
                        "Punto de Venta": "PV",
                        "Número de Comprobante": "Nro.",
                        "Nro. Doc. Vendedor": "CUIT",
                        "Denominación Vendedor": "Razón Social",
                        "Importe de Impuestos Internos": "Imp. Int.",
                        "Neto Gravado IVA 10,5%": "Gravado 10,5%",
                        "Neto Gravado IVA 21%": "Gravado 21%",
                        "Neto Gravado IVA 27%": "Gravado 27%",
                        "Importe IVA 10,5%": "IVA 10,5%",
                        "Importe IVA 21%": "IVA 21%",
                        "Importe IVA 27%": "IVA 27%",
                        "Importe Total": "Total",
                    }
                    # Verificar si las columnas existen antes de renombrarlas
                    existing_columns = {
                        old: new
                        for old, new in column_rename.items()
                        if old in df.columns
                    }
                    if existing_columns:
                        df = df.rename(columns=existing_columns)

                    # Equivalencias de comprobantes (mapeo)
                    if "Comprobante" in df.columns:

                        def code_to_str(val):
                            try:
                                return f"{int(float(str(val).replace(',', '.'))):03d}"
                            except:
                                return str(val)

                        df["Comprobante"] = (
                            df["Comprobante"]
                            .apply(code_to_str)
                            .map(comprobante_dict)
                            .fillna(df["Comprobante"])
                        )

                    # Eliminar columnas específicas
                    columns_to_remove = [
                        "Moneda Original",
                        "Tipo de Cambio",
                        "Total Neto Gravado",
                        "Total IVA",
                        "Importe Otros Tributos",
                        "Importe de Impuestos Municipales",
                        "Tipo Doc. Vendedor",
                        "Neto Gravado IVA 0%",
                        "Importe de Per. o Pagos a Cta. de Otros Imp. Nac.",
                        "Crédito Fiscal Computable",
                        "Neto Gravado IVA 2,5%",
                        "Importe IVA 2,5%",
                        "Neto Gravado IVA 5%",
                        "Importe IVA 5%",
                        "Exento",
                    ]
                    # Verificar si las columnas existen antes de eliminarlas
                    existing_columns = [
                        col for col in columns_to_remove if col in df.columns
                    ]
                    if existing_columns:
                        df = df.drop(columns=existing_columns)

                    # Reemplazar todos los None por "0,00" en todo el DataFrame
                    df = df.fillna("0,00")

                    # Convertir columnas a texto
                    text_columns = ["CUIT", "PV", "Nro."]
                    for col in text_columns:
                        if col in df.columns:
                            df[col] = df[col].astype(str)

                    # Limpiar y convertir PV, Nro. y CUIT a enteros puros
                    for col in ["PV", "Nro."]:
                        if col in df.columns:
                            df[col] = (
                                df[col]
                                .astype(str)
                                .str.replace(r"[^0-9]", "", regex=True)
                            )
                            df[col] = (
                                pd.to_numeric(df[col], errors="coerce")
                                .fillna(0)
                                .astype(int)
                            )

                    # Mover la columna 'Total' al final
                    if "Total" in df.columns:
                        cols = [c for c in df.columns if c != "Total"] + ["Total"]
                        df = df[cols]

                    # Eliminar puntos de miles y reemplazar coma decimal por punto en columnas de montos antes de convertir a float
                    if "No Gravado" in df.columns and "Total" in df.columns:
                        start = df.columns.get_loc("No Gravado")
                        end = df.columns.get_loc("Total")
                        for col in df.columns[start : end + 1]:
                            df[col] = (
                                df[col].astype(str).str.replace(".", "", regex=False)
                            )
                            df[col] = df[col].str.replace(",", ".", regex=False)
                            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(
                                0.0
                            )

                    # Ordenar el DataFrame por la columna 'Fecha' de menor a mayor si existe
                    if "Fecha" in df.columns:
                        df = df.sort_values(
                            "Fecha", ascending=True, na_position="last"
                        ).reset_index(drop=True)

                    # Limpieza previa y conversión robusta de la columna 'Fecha' (sin print)
                    if "Fecha" in df.columns:
                        df["Fecha"] = (
                            df["Fecha"]
                            .astype(str)
                            .str.strip()
                            .str.replace("-", "/", regex=False)
                        )
                        df["Fecha"] = pd.to_datetime(
                            df["Fecha"], dayfirst=True, errors="coerce"
                        )

                    # Vista previa
                    st.write("### Vista previa del archivo CSV:")
                    st.dataframe(df.head(), hide_index=True)

                    # Forzar CUIT a string antes de exportar
                    if "CUIT" in df.columns:
                        df["CUIT"] = df["CUIT"].astype(str)

                    # Extraer tipo ('Compras' o 'Ventas') y mes/año del nombre del archivo ZIP (patrón aaaaMM)
                    tipo_mes_anio = ""
                    if uploaded_file is not None:
                        zip_name = uploaded_file.name.lower()
                        tipo = (
                            "Compras"
                            if "compra" in zip_name
                            else ("Ventas" if "venta" in zip_name else "")
                        )
                        # Buscar patrón aaaaMM
                        match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", zip_name)
                        if match:
                            anio, mes = match.group(1), match.group(2)
                            tipo_mes_anio = (
                                f"{tipo} {mes}/{anio}" if tipo else f"{mes}/{anio}"
                            )
                        else:
                            tipo_mes_anio = tipo

                    # Exportación a Excel con nombre del contribuyente en dos filas, y celda fusionada con tipo y mes/año arriba de las columnas
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        # Escribir la tabla de datos a partir de la fila 5 (índice 4)
                        df.to_excel(writer, index=False, sheet_name="Datos", startrow=4)
                        ws = writer.sheets["Datos"]

                        # Nombre del contribuyente en las filas 1 y 2, fusionadas, en negrita y centrado, con borde grueso negro
                        ws.merge_cells(
                            start_row=1,
                            start_column=1,
                            end_row=2,
                            end_column=len(df.columns),
                        )
                        cell = ws.cell(row=1, column=1)
                        cell.value = contribuyente
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        thick_border = Border(
                            left=Side(style="thick", color="000000"),
                            right=Side(style="thick", color="000000"),
                            top=Side(style="thick", color="000000"),
                            bottom=Side(style="thick", color="000000"),
                        )
                        for row in range(1, 3):
                            for col in range(1, len(df.columns) + 1):
                                ws.cell(row=row, column=col).border = thick_border
                        # Fila 3 vacía como separador
                        # Celda fusionada con tipo y mes/año en la fila 4, arriba de los nombres de las columnas
                        if tipo_mes_anio:
                            ws.merge_cells(
                                start_row=4,
                                start_column=1,
                                end_row=4,
                                end_column=len(df.columns),
                            )
                            cell2 = ws.cell(row=4, column=1)
                            cell2.value = tipo_mes_anio
                            cell2.font = Font(bold=True)
                            cell2.alignment = Alignment(horizontal="center")
                        # Formato de moneda y formato condicional para negativos
                        if "No Gravado" in df.columns and "Total" in df.columns:
                            start = df.columns.get_loc("No Gravado") + 1  # Excel base 1
                            end = df.columns.get_loc("Total") + 1
                            from openpyxl.formatting.rule import CellIsRule

                            for col_idx in range(start, end + 1):
                                for row in range(5, 5 + len(df)):
                                    cell = ws.cell(row=row, column=col_idx)
                                    cell.number_format = '"$" #,##0.00'
                                col_letter = get_column_letter(col_idx)
                                ws.conditional_formatting.add(
                                    f"{col_letter}5:{col_letter}{len(df)+4}",
                                    CellIsRule(
                                        operator="lessThan",
                                        formula=["0"],
                                        font=Font(color="FF0000"),
                                    ),
                                )
                        # Formato de fecha para la columna 'Fecha'
                        if "Fecha" in df.columns:
                            fecha_idx = df.columns.get_loc("Fecha") + 1  # Excel base 1
                            for row in range(5, 5 + len(df)):
                                cell = ws.cell(row=row, column=fecha_idx)
                                try:
                                    if (
                                        isinstance(cell.value, str)
                                        and "-" in cell.value
                                    ):
                                        cell.value = pd.to_datetime(
                                            cell.value
                                        ).to_pydatetime()
                                except Exception:
                                    pass
                                cell.number_format = "DD/MM/YYYY"
                        # Formato de texto para la columna CUIT
                        if "CUIT" in df.columns:
                            from openpyxl.styles import numbers

                            cuit_idx = df.columns.get_loc("CUIT") + 1  # Excel base 1
                            for row in range(5, 5 + len(df)):
                                cell = ws.cell(row=row, column=cuit_idx)
                                cell.number_format = numbers.FORMAT_TEXT
                    output.seek(0)
                    st.download_button(
                        label="Descargar Excel",
                        data=output,
                        file_name="Datos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            else:
                st.error("No se encontraron archivos CSV en el archivo ZIP.")
    except Exception as e:
        st.error(f"Error al procesar el archivo: {str(e)}")
        st.write("Sugerencias para resolver el error:")
        st.write("1. Verifica que el archivo CSV esté correctamente formateado")
        st.write(
            "2. Asegúrate de que todas las filas tengan el mismo número de columnas"
        )
        st.write(
            "3. Si hay comas dentro de los campos, asegúrate de que estén entre comillas"
        )
